#!/usr/bin/env python3
"""
Kuaför Sepeti → Seyfibaba toplu ürün import Excel/CSV dönüştürücü.

Kullanım:
  python scripts/kuaforsepeti_scraper.py --limit 50
  python scripts/kuaforsepeti_scraper.py --all-categories --limit 0
  python scripts/kuaforsepeti_scraper.py --all-categories --limit 500
  python scripts/kuaforsepeti_scraper.py --category sac-boyalari --download-images

Çıktı: scripts/output/kuaforsepeti_import.xlsx (+ opsiyonel images/)
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import unicodedata
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

BASE_URL = "https://www.kuaforsepeti.com"
MARKUP = 1.20
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "tr-TR,tr;q=0.9,en;q=0.8",
}
OUTPUT_DIR = Path(__file__).resolve().parent / "output"
IMAGES_DIR = OUTPUT_DIR / "images"

# Ana kategoriler (siteden)
DEFAULT_CATEGORIES = [
    "sac-bakimi",
    "erkek-bakimi",
    "agda-ve-epilasyon-urunleri",
    "hijyen-urunleri",
    "el-ayak-ve-tirnak-bakimi",
    "cilt-bakimi",
    "vucut-bakimi",
    "demirbas-urunleri",
    "steril-cihazlari",
    "agiz-bakimi",
]

EXCEL_HEADERS = [
    "name",
    "short_name",
    "slug",
    "category",
    "sub_category",
    "child_category",
    "brand",
    "price",
    "offer_price",
    "qty",
    "short_description",
    "long_description",
    "sku",
    "weight",
    "tags",
    "image_url",
    "source_url",
    "source_price",
]


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9\s-]", "", text).lower()
    text = re.sub(r"[\s_-]+", "-", text).strip("-")
    return text or "urun"


def parse_price(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = (
        str(value)
        .replace("TL", "")
        .replace("₺", "")
        .replace(" ", "")
        .strip()
    )

    if cleaned == "":
        return None

    # 1.680,50 (TR binlik) veya 680,50
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", cleaned):
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    # aksi halde 680.00 gibi JSON fiyatına dokunma

    try:
        return float(cleaned)
    except ValueError:
        return None


def fetch_html(session: requests.Session, url: str) -> str:
    response = session.get(url, headers=HEADERS, timeout=45)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def extract_product_urls_from_category(html: str) -> list[str]:
    urls = set(re.findall(r'href="(/urun/[^"?#]+)"', html))
    urls.update(re.findall(r'href="(https://www\.kuaforsepeti\.com/urun/[^"?#]+)"', html))
    normalized = []
    for url in urls:
        if url.startswith("/"):
            url = urljoin(BASE_URL, url)
        normalized.append(url.split("?")[0])
    return sorted(set(normalized))


def collect_product_urls_from_sitemap(session: requests.Session) -> list[str]:
    """Sitemap'teki tüm ürün URL'lerini toplar (kategori taramasından daha eksiksiz)."""
    urls: set[str] = set()
    try:
        index_xml = fetch_html(session, f"{BASE_URL}/sitemap.xml")
    except requests.RequestException as exc:
        print(f"[warn] Sitemap index okunamadı: {exc}", file=sys.stderr)
        return []

    product_maps = re.findall(
        r"<loc>(https?://[^<]+/sitemap/products/\d+\.xml)</loc>",
        index_xml,
        flags=re.I,
    )
    if not product_maps:
        product_maps = [f"{BASE_URL}/sitemap/products/1.xml"]

    for map_url in product_maps:
        try:
            xml = fetch_html(session, map_url)
            for loc in re.findall(r"<loc>(https?://[^<]+/urun/[^<]+)</loc>", xml, flags=re.I):
                urls.add(loc.split("?")[0])
            time.sleep(0.1)
        except requests.RequestException as exc:
            print(f"[warn] Sitemap okunamadı {map_url}: {exc}", file=sys.stderr)

    return sorted(urls)


def extract_category_slug(href: str) -> str:
    if not href:
        return ""
    match = re.search(r"/kategori/([a-z0-9\-]+)", href, flags=re.I)
    return match.group(1).lower() if match else ""


def parse_breadcrumb_categories(soup: BeautifulSoup) -> list[tuple[str, str]]:
    """Breadcrumb'tan [(slug, ad), ...] — ilk eleman ana kategori."""
    for selector in (".breadcrumb", ".breadcrumb-wrapper", "ol.breadcrumb"):
        el = soup.select_one(selector)
        if not el:
            continue
        crumbs: list[tuple[str, str]] = []
        for link in el.select('a[href*="/kategori/"]'):
            slug = extract_category_slug(link.get("href", ""))
            name = link.get_text(strip=True)
            if slug and name:
                crumbs.append((slug, name))
        if crumbs:
            return crumbs
    return []


def breadcrumb_to_import_fields(crumbs: list[tuple[str, str]]) -> tuple[str, str, str]:
    if not crumbs:
        return "", "", ""
    category = crumbs[0][1]
    if len(crumbs) == 1:
        return category, "", ""
    if len(crumbs) == 2:
        return category, crumbs[1][1], ""
    return category, crumbs[1][1], crumbs[-1][1]


def register_breadcrumb_in_tree(tree: dict[str, dict], crumbs: list[tuple[str, str]]) -> None:
    if not crumbs:
        return
    main_slug, main_name = crumbs[0]
    tree[main_slug] = {
        "name": main_name,
        "parent_slug": None,
        "parent_name": None,
        "is_main": True,
    }
    for slug, name in crumbs[1:]:
        tree[slug] = {
            "name": name,
            "parent_slug": main_slug,
            "parent_name": main_name,
            "is_main": False,
        }


def load_category_breadcrumb(
    session: requests.Session,
    slug: str,
    cache: dict[str, list[tuple[str, str]]],
    tree: dict[str, dict],
) -> list[tuple[str, str]]:
    slug = slug.strip().lower()
    if not slug:
        return []
    if slug in cache:
        return cache[slug]
    try:
        html = fetch_html(session, f"{BASE_URL}/kategori/{slug}")
        soup = BeautifulSoup(html, "html.parser")
        crumbs = parse_breadcrumb_categories(soup)
        cache[slug] = crumbs
        register_breadcrumb_in_tree(tree, crumbs)
        time.sleep(0.15)
        return crumbs
    except requests.RequestException:
        cache[slug] = []
        return []


def build_category_tree(session: requests.Session) -> dict[str, dict]:
    """Ana kategorileri kaydeder; alt kategoriler ürün/breadcrumb ile tamamlanır."""
    tree: dict[str, dict] = {}

    html = fetch_html(session, BASE_URL)
    soup = BeautifulSoup(html, "html.parser")
    for link in soup.select('a[href*="/kategori/"]'):
        slug = extract_category_slug(link.get("href", ""))
        name = link.get_text(strip=True)
        if slug in DEFAULT_CATEGORIES and name:
            tree[slug] = {
                "name": name,
                "parent_slug": None,
                "parent_name": None,
                "is_main": True,
            }

    for main_slug in DEFAULT_CATEGORIES:
        if main_slug in tree:
            continue
        try:
            cat_html = fetch_html(session, f"{BASE_URL}/kategori/{main_slug}")
            cat_soup = BeautifulSoup(cat_html, "html.parser")
            crumbs = parse_breadcrumb_categories(cat_soup)
            if crumbs:
                register_breadcrumb_in_tree(tree, crumbs)
            else:
                h1 = cat_soup.find("h1")
                main_name = h1.get_text(strip=True) if h1 else main_slug.replace("-", " ").title()
                tree[main_slug] = {
                    "name": main_name,
                    "parent_slug": None,
                    "parent_name": None,
                    "is_main": True,
                }
            time.sleep(0.2)
        except requests.RequestException:
            tree[main_slug] = {
                "name": main_slug.replace("-", " ").title(),
                "parent_slug": None,
                "parent_name": None,
                "is_main": True,
            }

    return tree


def resolve_import_categories(
    tree: dict[str, dict],
    slug: str = "",
    name: str = "",
    session: requests.Session | None = None,
    breadcrumb_cache: dict[str, list[tuple[str, str]]] | None = None,
) -> tuple[str, str, str]:
    slug = (slug or "").strip().lower()
    name = (name or "").strip()

    if slug and session is not None:
        cache = breadcrumb_cache if breadcrumb_cache is not None else {}
        crumbs = load_category_breadcrumb(session, slug, cache, tree)
        if crumbs:
            return breadcrumb_to_import_fields(crumbs)

    if slug and slug in tree:
        node = tree[slug]
        if node.get("is_main"):
            return node["name"], "", ""
        parent_name = node.get("parent_name") or ""
        if parent_name:
            return parent_name, node["name"], ""
        return node["name"], "", ""

    if name:
        lowered = name.casefold()
        for s, node in tree.items():
            if node["name"].casefold() == lowered:
                return resolve_import_categories(
                    tree,
                    slug=s,
                    name=node["name"],
                    session=session,
                    breadcrumb_cache=breadcrumb_cache,
                )

    if name:
        return "", name, ""
    return "", "", ""


def discover_category_slugs(session: requests.Session, tree: dict[str, dict] | None = None) -> list[str]:
    html = fetch_html(session, BASE_URL)
    slugs = set(tree.keys() if tree else [])
    slugs.update(re.findall(r"/kategori/([a-z0-9\-]+)", html))
    slugs.update(DEFAULT_CATEGORIES)

    # Alt kategori sayfalarından ek slug topla
    for slug in list(slugs)[:80]:
        try:
            cat_html = fetch_html(session, f"{BASE_URL}/kategori/{slug}")
            slugs.update(re.findall(r"/kategori/([a-z0-9\-]+)", cat_html))
            time.sleep(0.2)
        except requests.RequestException:
            pass

    return sorted(slugs)


def collect_product_urls(
    session: requests.Session,
    category_slugs: list[str],
    category_tree: dict[str, dict],
    breadcrumb_cache: dict[str, list[tuple[str, str]]],
    max_pages_per_category: int = 50,
) -> tuple[list[str], dict[str, dict]]:
    all_urls: set[str] = set()
    url_categories: dict[str, dict] = {}

    for slug in category_slugs:
        empty_streak = 0
        listing_name = slug.replace("-", " ").title()
        listing_category, listing_sub, listing_child = resolve_import_categories(
            category_tree, slug, listing_name, session=session, breadcrumb_cache=breadcrumb_cache
        )
        for page in range(1, max_pages_per_category + 1):
            url = f"{BASE_URL}/kategori/{slug}?sayfa={page}" if page > 1 else f"{BASE_URL}/kategori/{slug}"
            try:
                html = fetch_html(session, url)
            except requests.RequestException as exc:
                print(f"[warn] Kategori sayfası alınamadı {url}: {exc}", file=sys.stderr)
                break

            soup = BeautifulSoup(html, "html.parser")
            heading = soup.find("h1")
            if heading and heading.get_text(strip=True):
                listing_name = heading.get_text(strip=True)
                listing_category, listing_sub, listing_child = resolve_import_categories(
                    category_tree,
                    slug,
                    listing_name,
                    session=session,
                    breadcrumb_cache=breadcrumb_cache,
                )

            page_urls = extract_product_urls_from_category(html)
            new_urls = [u for u in page_urls if u not in all_urls]
            ctx = {
                "category": listing_category,
                "sub_category": listing_sub,
                "child_category": listing_child,
                "listing_slug": slug,
            }
            for product_url in page_urls:
                existing = url_categories.get(product_url)
                if not existing or (ctx["sub_category"] and not existing.get("sub_category")):
                    url_categories[product_url] = ctx
            all_urls.update(page_urls)

            print(f"  {slug} sayfa {page}: +{len(new_urls)} ürün (toplam {len(all_urls)})")

            if not new_urls:
                empty_streak += 1
                if empty_streak >= 2:
                    break
            else:
                empty_streak = 0

            time.sleep(0.35)

    return sorted(all_urls), url_categories


def parse_product_json_ld(html: str) -> dict | None:
    for match in re.finditer(
        r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>',
        html,
        flags=re.S | re.I,
    ):
        raw = match.group(1).strip()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict) and data.get("@type") == "Product":
            return data
    return None


def parse_product_category(soup: BeautifulSoup) -> tuple[str, str]:
    """Ürün sayfasından (slug, görünen ad) döner — genelde alt kategori."""
    for li in soup.find_all("li"):
        text = li.get_text(" ", strip=True)
        if not text.lower().startswith("kategori:"):
            continue
        link = li.find("a", href=re.compile(r"/kategori/"))
        if link:
            return extract_category_slug(link.get("href", "")), link.get_text(strip=True)

    for text_node in soup.find_all(string=re.compile(r"^\s*Kategori\s*:\s*$", re.I)):
        container = text_node.parent
        if not container:
            continue
        link = container.find("a", href=re.compile(r"/kategori/"))
        if not link and container.parent:
            link = container.parent.find("a", href=re.compile(r"/kategori/"))
        if link:
            return extract_category_slug(link.get("href", "")), link.get_text(strip=True)

    return "", ""


def parse_product_meta(soup: BeautifulSoup) -> dict:
    meta = {}
    category_slug, category_name = parse_product_category(soup)
    if category_name:
        meta["category"] = category_name
        meta["category_slug"] = category_slug

    for label in soup.select(".product-detail, .product-profile, .product-info, .detail"):
        text = label.get_text("\n", strip=True)
        if "Marka:" in text and "brand" not in meta:
            link = label.select_one('a[href*="/marka/"]')
            if link:
                meta["brand"] = link.get_text(strip=True)
        if "Ürün Kodu:" in text:
            m = re.search(r"Ürün Kodu:\s*(.+)", text)
            if m:
                meta["product_code"] = m.group(1).strip()

    og_desc = soup.find("meta", attrs={"property": "og:description"})
    if og_desc and og_desc.get("content"):
        meta["description"] = og_desc["content"].strip()

    keywords = soup.find("meta", attrs={"property": "og:keywords"})
    if keywords and keywords.get("content"):
        meta["tags"] = keywords["content"].strip()

    return meta


def best_image_url(images) -> str:
    if isinstance(images, str):
        return images
    if isinstance(images, list) and images:
        return images[0]
    return ""


def high_res_image_url(url: str) -> str:
    if not url:
        return ""
    # JSON-LD / og:image URL'lerini olduğu gibi kullan (CDN dönüşümü 404 verebilir)
    return url.strip()


def scrape_product(
    session: requests.Session,
    url: str,
    category_tree: dict[str, dict],
    fallback_context: dict | str | None = None,
    breadcrumb_cache: dict[str, list[tuple[str, str]]] | None = None,
) -> dict | None:
    try:
        html = fetch_html(session, url)
    except requests.RequestException as exc:
        print(f"[warn] Ürün alınamadı {url}: {exc}", file=sys.stderr)
        return None

    product_ld = parse_product_json_ld(html)
    soup = BeautifulSoup(html, "html.parser")
    meta = parse_product_meta(soup)

    name = (product_ld or {}).get("name") or (soup.find("h1").get_text(strip=True) if soup.find("h1") else "")
    if not name:
        return None

    offers = (product_ld or {}).get("offers") or {}
    if isinstance(offers, list):
        offers = offers[0] if offers else {}

    source_price = parse_price(offers.get("price"))
    if source_price is None:
        price_match = re.search(r"([\d\.]+,\d{2})\s*TL", html)
        if price_match:
            source_price = parse_price(price_match.group(1))

    if source_price is None:
        print(f"[warn] Fiyat bulunamadı: {url}", file=sys.stderr)
        return None

    marked_up = round(source_price * MARKUP, 2)

    brand = ""
    if product_ld and product_ld.get("brand"):
        brand_data = product_ld["brand"]
        brand = brand_data.get("name") if isinstance(brand_data, dict) else str(brand_data)
    brand = brand or meta.get("brand", "")

    sku = (product_ld or {}).get("sku") or (product_ld or {}).get("mpn") or meta.get("product_code") or ""
    description = (product_ld or {}).get("description") or meta.get("description") or name
    tags = meta.get("tags", "")

    cat_slug, cat_name = parse_product_category(soup)
    if not cat_slug:
        cat_slug = meta.get("category_slug", "")
    if not cat_name:
        cat_name = meta.get("category", "")

    category, sub_category, child_category = resolve_import_categories(
        category_tree, cat_slug, cat_name, session=session, breadcrumb_cache=breadcrumb_cache
    )

    if not category and isinstance(fallback_context, dict):
        category = fallback_context.get("category") or category
        sub_category = fallback_context.get("sub_category") or sub_category
        child_category = fallback_context.get("child_category") or child_category
    elif not category and isinstance(fallback_context, str) and fallback_context:
        category, sub_category, child_category = resolve_import_categories(
            category_tree, name=fallback_context, session=session, breadcrumb_cache=breadcrumb_cache
        )

    if not category and not sub_category and cat_slug:
        category, sub_category, child_category = resolve_import_categories(
            category_tree, cat_slug, cat_name, session=session, breadcrumb_cache=breadcrumb_cache
        )

    if category and not sub_category and cat_name and cat_name.casefold() != category.casefold():
        sub_category = cat_name

    image_url = high_res_image_url(best_image_url((product_ld or {}).get("image")))

    short_name = name[:60] + ("..." if len(name) > 60 else "")

    return {
        "name": name,
        "short_name": short_name,
        "slug": slugify(name),
        "category": category,
        "sub_category": sub_category,
        "child_category": child_category,
        "brand": brand,
        "price": f"{marked_up:.2f}",
        "offer_price": "",
        "qty": "10",
        "short_description": description[:250] if description else name[:250],
        "long_description": description,
        "sku": sku,
        "weight": "",
        "tags": tags.replace(";", ","),
        "image_url": image_url,
        "source_url": url,
        "source_price": f"{source_price:.2f}",
    }


def download_image(session: requests.Session, url: str, dest: Path) -> str:
    if not url or dest.exists():
        return str(dest) if dest.exists() else url

    try:
        response = session.get(url, headers=HEADERS, timeout=60)
        response.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(response.content)
        return str(dest)
    except requests.RequestException as exc:
        print(f"[warn] Görsel indirilemedi {url}: {exc}", file=sys.stderr)
        return url


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Urunler"
    ws.append(EXCEL_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in EXCEL_HEADERS])

    guide = wb.create_sheet("Nasil Kullanilir")
    guide.append(["Seyfibaba Toplu Ürün Yükleme — Kuaför Sepeti aktarım dosyası"])
    guide.append([f"Fiyatlara %{int((MARKUP - 1) * 100)} marj uygulandı (source_price = kaynak fiyat)."])
    guide.append(["Görsel URL'leri CDN üzerinden; import sırasında sistemin indirmesi gerekir."])
    guide.append(["Zorunlu: name, category, price, qty"])
    guide.append(["category = ana kategori, sub_category = alt kategori (Seyfibaba import eşleşmesi)"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXCEL_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def repair_categories_in_file(
    session: requests.Session,
    input_path: Path,
    output_path: Path,
    category_tree: dict[str, dict],
) -> int:
    """Mevcut Excel/CSV'deki ürünlerin kategori alanlarını ürün sayfasından yeniden doldurur."""
    rows: list[dict] = []
    suffix = input_path.suffix.lower()

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(input_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    elif suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
    else:
        raise ValueError("Desteklenen formatlar: .xlsx, .csv")

    breadcrumb_cache: dict[str, list[tuple[str, str]]] = {}
    updated = 0
    for index, row in enumerate(rows, start=1):
        source_url = (row.get("source_url") or "").strip()
        if not source_url:
            continue
        try:
            html = fetch_html(session, source_url)
        except requests.RequestException as exc:
            print(f"[warn] [{index}] Kategori okunamadı: {exc}", file=sys.stderr)
            continue
        soup = BeautifulSoup(html, "html.parser")
        cat_slug, cat_name = parse_product_category(soup)
        category, sub_category, child_category = resolve_import_categories(
            category_tree,
            cat_slug,
            cat_name,
            session=session,
            breadcrumb_cache=breadcrumb_cache,
        )
        if category or sub_category:
            if category and not sub_category and cat_name and cat_name.casefold() != category.casefold():
                sub_category = cat_name
            row["category"] = category
            row["sub_category"] = sub_category
            row["child_category"] = child_category
            updated += 1
            suffix = f" / {child_category}" if child_category else ""
            print(f"[{index}/{len(rows)}] {row.get('name', '')[:50]} -> {category} / {sub_category}{suffix}")
        time.sleep(0.35)

    if suffix == ".xlsx":
        write_xlsx(rows, output_path)
        write_csv(rows, output_path.with_suffix(".csv"))
    else:
        write_csv(rows, output_path)
        write_xlsx(rows, output_path.with_suffix(".xlsx"))

    return updated


def main() -> int:
    parser = argparse.ArgumentParser(description="Kuaför Sepeti ürün scraper → Seyfibaba Excel")
    parser.add_argument("--category", action="append", help="Tek kategori slug (ör. sac-boyalari)")
    parser.add_argument("--all-categories", action="store_true", help="Tüm ana kategorileri tara")
    parser.add_argument("--urls-file", help="Ürün URL listesi (satır başına bir URL)")
    parser.add_argument("--limit", type=int, default=100, help="Maksimum ürün sayısı (0 = sınırsız)")
    parser.add_argument("--download-images", action="store_true", help="Görselleri yerel klasöre indir")
    parser.add_argument("--repair-categories", help="Mevcut xlsx/csv dosyasındaki kategori alanlarını ürün sayfasından düzelt")
    parser.add_argument("--output", default=str(OUTPUT_DIR / "kuaforsepeti_import.xlsx"))
    args = parser.parse_args()

    session = requests.Session()
    breadcrumb_cache: dict[str, list[tuple[str, str]]] = {}

    print("Kategori ağacı oluşturuluyor...")
    category_tree = build_category_tree(session)
    print(f"  {len(category_tree)} kategori/alt kategori kaydı")

    if args.repair_categories:
        input_path = Path(args.repair_categories)
        output_path = Path(args.output)
        count = repair_categories_in_file(session, input_path, output_path, category_tree)
        print(f"\nTamamlandı: {count} ürünün kategori alanı güncellendi")
        print(f"Excel: {output_path}")
        print(f"CSV:   {output_path.with_suffix('.csv')}")
        return 0 if count else 1

    product_urls: list[str] = []
    url_categories: dict[str, dict] = {}

    if args.urls_file:
        product_urls = [
            line.strip()
            for line in Path(args.urls_file).read_text(encoding="utf-8").splitlines()
            if line.strip().startswith("http")
        ]
    else:
        if args.category:
            categories = args.category
        elif args.all_categories:
            print("Sitemap'ten ürün URL'leri toplanıyor...")
            product_urls = collect_product_urls_from_sitemap(session)
            print(f"  {len(product_urls)} ürün URL'si (sitemap)")
            print("Kategori bağlamı için liste sayfaları taranıyor...")
            categories = discover_category_slugs(session, category_tree)
            _, url_categories = collect_product_urls(
                session, categories, category_tree, breadcrumb_cache
            )
        else:
            categories = DEFAULT_CATEGORIES[:3]  # varsayılan: ilk 3 ana kategori

        if not product_urls:
            print(f"Kategoriler taranıyor ({len(categories)} adet)...")
            product_urls, url_categories = collect_product_urls(
                session, categories, category_tree, breadcrumb_cache
            )

    if args.limit > 0:
        product_urls = product_urls[: args.limit]

    print(f"\n{len(product_urls)} ürün detayı çekiliyor...\n")

    rows: list[dict] = []
    seen_slugs: set[str] = set()

    for index, url in enumerate(product_urls, start=1):
        row = scrape_product(
            session, url, category_tree, url_categories.get(url), breadcrumb_cache
        )
        if not row:
            continue

        base_slug = row["slug"]
        slug = base_slug
        counter = 2
        while slug in seen_slugs:
            slug = f"{base_slug}-{counter}"
            counter += 1
        row["slug"] = slug
        seen_slugs.add(slug)

        if args.download_images and row.get("image_url"):
            ext = Path(urlparse(row["image_url"]).path).suffix or ".webp"
            image_path = IMAGES_DIR / f"{slug}{ext}"
            local_path = download_image(session, row["image_url"], image_path)
            row["image_url"] = local_path

        rows.append(row)
        print(f"[{index}/{len(product_urls)}] {row['name'][:70]} -> {row['price']} TL")

        time.sleep(0.4)

    if not rows:
        print("Hiç ürün çekilemedi.", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    write_xlsx(rows, output_path)
    write_csv(rows, output_path.with_suffix(".csv"))

    print(f"\nTamamlandı: {len(rows)} ürün")
    print(f"Excel: {output_path}")
    print(f"CSV:   {output_path.with_suffix('.csv')}")
    if args.download_images:
        print(f"Görseller: {IMAGES_DIR}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
